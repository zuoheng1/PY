import datetime
import os
from openpyxl import load_workbook
import shutil
import zipfile

start_year = 2025  #开始年份
satrt_month = 2   #开始的月份
filesNum = 29     #设置需要的子文件数量


original_folder_path = '/Users/anker/Desktop/py'  
template_path = '/Users/anker/Desktop/py/Template.xlsx' 
generated_folder_path = '/Users/anker/Desktop/py/GeneratedExcelFiles' 
zip_file_path =  f'/Users/anker/Desktop/py/{start_year}年_{satrt_month}月份计划子表汇总.zip' 
file_name = os.path.basename(zip_file_path)
start_date = datetime.date(start_year, satrt_month, 1)  


if not os.path.exists(original_folder_path):
    os.makedirs(original_folder_path)
if not os.path.exists(generated_folder_path):
    os.makedirs(generated_folder_path)


template_wb = load_workbook(template_path)
template_ws = template_wb.active

def generate_excel_for_days(start_date, filesNum, generated_folder_path, template_wb):
    for i in range(filesNum):
        current_date = start_date + datetime.timedelta(days=i)
        filename = current_date.strftime('%Y-%m-%d') + "计划完成率" + '.xlsx'
        file_path = os.path.join(generated_folder_path, filename)
        
        new_wb = load_workbook(template_path)
        
        new_wb.save(file_path)

        new_wb.close()

def move_and_compress_files(generated_folder_path, zip_file_path):
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(generated_folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                in_zip_path = os.path.relpath(file_path, generated_folder_path)
                zipf.write(file_path, in_zip_path)

    for file_name in os.listdir(generated_folder_path):
        file_path = os.path.join(generated_folder_path, file_name)
        os.remove(file_path)

generate_excel_for_days(start_date, filesNum, generated_folder_path, template_wb)

move_and_compress_files(generated_folder_path, zip_file_path)

print(f"All files have been compressed into zip: {zip_file_path}")

if os.path.exists(generated_folder_path):
    shutil.rmtree(generated_folder_path)
    print(f"{file_name}已创建，任务完成✅")