import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import deque
import re
from datetime import datetime
import os
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import json

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Link404Crawler:
    def __init__(self, domain, max_pages=100, delay=1, path_filter=None, 
                 max_workers=5, timeout=10):
        self.domain = domain
        self.base_url = self._normalize_url(domain)
        self.max_pages = max_pages
        self.delay = delay
        self.path_filter = path_filter
        self.timeout = timeout
        self.max_workers = max_workers
        
        # 数据存储
        self.visited_urls = set()
        self.found_404s = []
        self.all_links = set()
        self.page_link_details = []
        self._lock = threading.Lock()
        
        # HTTP会话配置
        self.session = self._create_session()
        
        # 静态资源文件扩展名列表
        self.static_extensions = {
            '.css', '.scss', '.sass', '.less',
            '.js', '.jsx', '.ts', '.tsx', '.coffee',
            '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp', '.ico', '.tiff', '.tif',
            '.woff', '.woff2', '.ttf', '.otf', '.eot',
            '.mp3', '.mp4', '.avi', '.mov', '.wmv', '.flv', '.webm', '.ogg', '.wav',
            '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.rtf',
            '.zip', '.rar', '.7z', '.tar', '.gz', '.bz2',
            '.xml', '.json', '.csv', '.swf', '.map'
        }
    
    def _normalize_url(self, domain):
        """标准化URL格式"""
        if not domain.startswith(('http://', 'https://')):
            return f"https://{domain}"
        return domain
    
    def _create_session(self):
        """创建HTTP会话"""
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        # 设置连接池大小
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=20,
            pool_maxsize=20,
            max_retries=3
        )
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        return session
    
    def is_static_resource(self, url):
        """检查URL是否为静态资源文件"""
        try:
            parsed = urlparse(url)
            path = parsed.path.lower()
            
            # 检查文件扩展名
            if any(path.endswith(ext) for ext in self.static_extensions):
                return True
            
            # 检查路径模式
            static_patterns = [
                '/_next/static/', '/static/', '/assets/', '/public/',
                '/dist/', '/build/', '/css/', '/js/', '/images/',
                '/img/', '/fonts/', '/media/'
            ]
            
            return any(pattern in path for pattern in static_patterns)
            
        except Exception:
            return False
    
    def is_valid_url(self, url):
        """检查URL是否有效且属于目标域名"""
        try:
            parsed = urlparse(url)
            base_parsed = urlparse(self.base_url)
            
            # 检查域名
            if parsed.netloc and parsed.netloc != base_parsed.netloc:
                return False
            
            # 检查协议
            if parsed.scheme in ['mailto', 'tel', 'javascript']:
                return False
            
            # 检查锚点
            if url.startswith('#'):
                return False
            
            # 检查静态资源
            if self.is_static_resource(url):
                return False
            
            return True
            
        except Exception:
            return False
    
    def matches_path_filter(self, url):
        """检查URL是否匹配路径过滤器"""
        if not self.path_filter:
            return True
        
        try:
            parsed = urlparse(url)
            path = parsed.path
            
            if isinstance(self.path_filter, list):
                return any(path.startswith(filter_path) for filter_path in self.path_filter)
            else:
                # 🔧 针对 /au 的精确匹配
                filter_path = self.path_filter
                # 确保路径以 / 开头
                if not filter_path.startswith('/'):
                    filter_path = '/' + filter_path
                
                # 检查路径是否以筛选条件开头
                return path.startswith(filter_path)
                
        except Exception as e:
            logger.error(f"路径匹配检查出错: {e}")
            return False
    
    def detect_link_position_and_classes(self, element):
        """检测链接在页面中的位置和class属性"""
        positions = []
        classes_info = []
        current = element
        
        position_indicators = {
            'header': ['header', 'top', 'navbar', 'nav-bar', 'navigation'],
            'footer': ['footer', 'bottom', 'foot'],
            'sidebar': ['sidebar', 'side-bar', 'aside'],
            'main': ['main', 'content', 'body'],
            'menu': ['menu', 'nav', 'navigation'],
            'breadcrumb': ['breadcrumb', 'breadcrumbs'],
            'pagination': ['pagination', 'pager'],
            'search': ['search', 'search-box'],
            'social': ['social', 'share', 'follow'],
            'product': ['product', 'item', 'card'],
            'category': ['category', 'cat', 'section'],
            'banner': ['banner', 'hero', 'slider'],
            'form': ['form', 'contact', 'subscribe']
        }
        
        # 收集当前元素的class信息
        element_classes = element.get('class', [])
        if element_classes:
            classes_info.append({
                'tag': element.name,
                'classes': ' '.join(element_classes),
                'level': 'current'
            })
        
        level = 0
        while current and current.name and level < 5:  # 最多向上查找5层
            # 检查HTML标签
            if current.name in ['header', 'footer', 'nav', 'aside', 'main', 'section', 'article']:
                positions.append(f"<{current.name}>")
            
            # 检查class和id属性
            classes = current.get('class', [])
            element_id = current.get('id', '')
            
            # 收集父级元素的class信息
            if classes and level > 0:  # 不重复收集当前元素
                classes_info.append({
                    'tag': current.name,
                    'classes': ' '.join(classes),
                    'level': f'parent-{level}'
                })
            
            all_attrs = ' '.join(classes + [element_id]).lower()
            
            for position, keywords in position_indicators.items():
                if any(keyword in all_attrs for keyword in keywords):
                    positions.append(f"{position}({keywords[0]})")
                    break
            
            current = current.parent
            level += 1
        
        position_str = " > ".join(reversed(positions)) if positions else "页面主体"
        
        # 生成CSS选择器
        css_selector = self._generate_css_selector(element)
        
        # 生成XPath
        xpath = self._generate_xpath(element)
        
        # 确定可视化位置
        visual_position = self._determine_visual_position(positions, element_classes)
        
        return {
            'position': position_str,
            'classes_info': classes_info,
            'element_id': element.get('id', ''),
            'element_tag': element.name,
            'css_selector': css_selector,
            'xpath': xpath,
            'visual_position': visual_position
        }
    
    def check_url_status(self, url):
        """检查URL的状态码"""
        try:
            # 先尝试HEAD请求
            response = self.session.head(url, timeout=self.timeout, allow_redirects=True)
            return response.status_code
        except requests.exceptions.RequestException:
            try:
                # HEAD失败则尝试GET请求
                response = self.session.get(url, timeout=self.timeout, allow_redirects=True)
                return response.status_code
            except Exception as e:
                logger.warning(f"检查URL状态失败 {url}: {e}")
                return 'ERROR'
    
    def check_urls_batch(self, urls):
        """批量检查URL状态"""
        results = {}
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_url = {executor.submit(self.check_url_status, url): url for url in urls}
            
            for future in as_completed(future_to_url):
                url = future_to_url[future]
                try:
                    status = future.result()
                    results[url] = status
                except Exception as e:
                    logger.error(f"检查URL {url} 时出错: {e}")
                    results[url] = 'ERROR'
        
        return results
    
    def extract_and_check_links_from_page(self, url):
        """从页面提取并检查链接"""
        try:
            logger.info(f"🔍 开始检测页面: {url}")
            response = self.session.get(url, timeout=self.timeout)
            
            if response.status_code != 200:
                logger.warning(f"⚠️ 页面访问失败: {url} (状态码: {response.status_code})")
                return set()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            links = set()
            link_positions = {}
            
            # 添加调试信息：显示页面基本信息
            logger.info(f"📄 页面标题: {soup.title.string if soup.title else '无标题'}")
            
            # 提取所有链接
            self._extract_links_from_soup(soup, url, links, link_positions)
            
            # 添加调试信息：显示提取到的链接数量
            logger.info(f"🔗 从页面提取到 {len(links)} 个原始链接")
            
            # 过滤有效链接
            valid_links = set()
            for link in links:
                if self.is_valid_url(link) and not self.is_static_resource(link):
                    valid_links.add(link)
            
            logger.info(f"✅ 过滤后有效链接: {len(valid_links)} 个")
            
            # 显示前几个链接作为样本
            if valid_links:
                logger.info(f"📋 链接样本 (前5个):")
                for i, link in enumerate(list(valid_links)[:5]):
                    logger.info(f"  {i+1}. {link}")
            
            page_links_status = []
            
            # 批量检查链接状态
            if links:
                logger.info(f"🔍 检查 {len(links)} 个链接的状态...")
                status_results = self.check_urls_batch(links)
                
                # 处理结果
                for link_url in links:
                    status = status_results.get(link_url, 'ERROR')
                    position_info = link_positions.get(link_url, {
                        'position': '未知位置', 
                        'text': '', 
                        'element_type': 'unknown',
                        'classes_info': [],
                        'element_id': '',
                        'element_tag': '',
                        'css_selector': '',
                        'xpath': '',
                        'visual_position': '未知区域'
                    })
                    
                    link_status = self._create_link_status(url, link_url, status, position_info)
                    page_links_status.append(link_status)
                    
                    # 处理404链接
                    if status == 404:
                        self._handle_404_link(url, link_url, position_info, link_status)
            
            # 保存页面详情
            self._save_page_details(url, links, page_links_status)
            
            return valid_links
            
        except Exception as e:
            logger.error(f"提取链接时出错 {url}: {e}")
            return set()
    
    def _extract_links_from_soup(self, soup, base_url, links, link_positions):
        """从BeautifulSoup对象中提取链接"""
        # 提取a标签链接
        for link in soup.find_all('a', href=True):
            href = link['href'].strip()
            if href:
                absolute_url = urljoin(base_url, href)
                if self.is_valid_url(absolute_url):
                    clean_url = absolute_url.split('#')[0]
                    links.add(clean_url)
                    if clean_url not in link_positions:
                        position_data = self.detect_link_position_and_classes(link)
                        link_positions[clean_url] = {
                            'position': position_data['position'],
                            'text': link.get_text(strip=True)[:50],
                            'element_type': 'link',
                            'classes_info': position_data['classes_info'],
                            'element_id': position_data['element_id'],
                            'element_tag': position_data['element_tag'],
                            'css_selector': position_data['css_selector'],
                            'xpath': position_data['xpath'],
                            'visual_position': position_data['visual_position']
                        }
        
        # 提取img标签链接
        for img in soup.find_all('img', src=True):
            src = img['src'].strip()
            if src:
                absolute_url = urljoin(base_url, src)
                if self.is_valid_url(absolute_url):
                    links.add(absolute_url)
                    if absolute_url not in link_positions:
                        position_data = self.detect_link_position_and_classes(img)
                        link_positions[absolute_url] = {
                            'position': position_data['position'],
                            'text': img.get('alt', '')[:50],
                            'element_type': 'image',
                            'classes_info': position_data['classes_info'],
                            'element_id': position_data['element_id'],
                            'element_tag': position_data['element_tag'],
                            'css_selector': position_data['css_selector'],
                            'xpath': position_data['xpath'],
                            'visual_position': position_data['visual_position']
                        }
    
    def _create_link_status(self, parent_url, link_url, status, position_info):
        """创建链接状态对象"""
        return {
            'parent_page': parent_url,
            'link_url': link_url,
            'status_code': status,
            'check_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'matches_filter': self.matches_path_filter(link_url),
            'position': position_info['position'],
            'link_text': position_info['text'],
            'element_type': position_info['element_type'],
            'classes_info': position_info.get('classes_info', []),
            'element_id': position_info.get('element_id', ''),
            'element_tag': position_info.get('element_tag', ''),
            'css_selector': position_info.get('css_selector', ''),
            'xpath': position_info.get('xpath', ''),
            'visual_position': position_info.get('visual_position', '')
        }
    
    def _handle_404_link(self, parent_url, link_url, position_info, link_status):
        """处理404链接"""
        filter_indicator = "🎯" if self.matches_path_filter(link_url) else "⚪"
        logger.info(f"    {filter_indicator} ❌ 404: {link_url}")
        logger.info(f"         📍 位置: [{position_info['visual_position']}]")
        logger.info(f"         🎯 CSS选择器: {position_info.get('css_selector', '未生成')}")
        
        if position_info['text']:
            logger.info(f"         📝 文本: {position_info['text']}")
        
        # 显示class信息
        if position_info.get('classes_info'):
            logger.info(f"         🏷️  Class信息:")
            for class_info in position_info['classes_info']:
                logger.info(f"             {class_info['level']}: <{class_info['tag']}> class=\"{class_info['classes']}\"")
        
        if position_info.get('element_id'):
            logger.info(f"         🆔 元素ID: {position_info['element_id']}")
        
        # 生成修复建议
        fix_suggestion = self.generate_fix_suggestion({
            'url': link_url,
            'visual_position': position_info.get('visual_position', '')
        })
        logger.info(f"         💡 修复建议: {fix_suggestion}")
        
        # 添加到404列表
        with self._lock:
            self.found_404s.append({
                'url': link_url,
                'found_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'status_code': 404,
                'parent_page': parent_url,
                'matches_filter': self.matches_path_filter(link_url),
                'position': position_info['position'],
                'link_text': position_info['text'],
                'element_type': position_info['element_type'],
                'classes_info': position_info.get('classes_info', []),
                'element_id': position_info.get('element_id', ''),
                'element_tag': position_info.get('element_tag', ''),
                'css_selector': position_info.get('css_selector', ''),
                'xpath': position_info.get('xpath', ''),
                'visual_position': position_info.get('visual_position', ''),
                'fix_suggestion': fix_suggestion
            })
    
    def _save_page_details(self, url, links, page_links_status):
        """保存页面详情"""
        with self._lock:
            self.page_link_details.append({
                'page_url': url,
                'total_links': len(links),
                'links_status': page_links_status,
                'scan_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # 统计404链接
        status_404_count = sum(1 for link_status in page_links_status 
                              if link_status['status_code'] == 404)
        
        if status_404_count > 0:
            logger.info(f"  📊 发现 {status_404_count} 个404链接")
    
    def print_final_summary(self):
        """打印最终统计摘要"""
        total_404s = len(self.found_404s)
        filtered_404s = sum(1 for link in self.found_404s if link['matches_filter'])
        total_pages = len(self.visited_urls)
        total_links = len(self.all_links)
        
        logger.info("\n" + "=" * 80)
        logger.info("🎯 检测完成 - 最终统计")
        logger.info("=" * 80)
        logger.info(f"📄 已检测页面: {total_pages}")
        logger.info(f"🔗 发现链接总数: {total_links}")
        logger.info(f"❌ 404链接总数: {total_404s}")
        logger.info(f"🎯 符合筛选条件的404链接: {filtered_404s}")
        
        if total_404s > 0:
            # 按位置分组统计
            position_stats = {}
            for link in self.found_404s:
                pos = link.get('visual_position', '未知区域')
                position_stats[pos] = position_stats.get(pos, 0) + 1
            
            logger.info("\n📊 404链接位置分布:")
            for position, count in sorted(position_stats.items(), key=lambda x: x[1], reverse=True):
                logger.info(f"  {position}: {count} 个")
        
        logger.info("=" * 80)
    
    def crawl_for_404s(self, start_url=None):
        """爬取域名下的所有链接并检测404"""
        if start_url is None:
            start_url = self._get_start_url()
        
        self._print_crawl_info(start_url)
        
        # 🔧 针对 /au 路径的特殊处理
        if self.path_filter and not start_url.endswith(self.path_filter.lstrip('/')):
            # 如果起始URL不包含筛选路径，自动添加
            parsed = urlparse(start_url)
            start_url = f"{parsed.scheme}://{parsed.netloc}{self.path_filter}"
            logger.info(f"🎯 自动调整起始URL为: {start_url}")
        
        url_queue = deque([start_url])
        pages_crawled = 0
        
        while url_queue and pages_crawled < self.max_pages:
            current_url = url_queue.popleft()
            
            if current_url in self.visited_urls:
                continue
            
            # 🔧 关键优化：在处理前就检查路径筛选
            if self.path_filter and not self.matches_path_filter(current_url):
                logger.info(f"⏭️  跳过不符合筛选条件的URL: {current_url}")
                continue
            
            self.visited_urls.add(current_url)
            pages_crawled += 1
            
            logger.info(f"🎯 当前队列长度: {len(url_queue)}, 已访问页面: {len(self.visited_urls)}")
            logger.info(f"\n📖 正在爬取第 {pages_crawled}/{self.max_pages} 页: {current_url}")
            
            status = self.check_url_status(current_url)
            
            if status == 404:
                self._handle_404_page(current_url)
            elif status == 'ERROR':
                logger.warning(f"⚠️  页面状态: 访问错误")
            elif status == 200:
                # 提取并检查页面链接
                links = self.extract_and_check_links_from_page(current_url)
                self.all_links.update(links)
                
                # 🔧 只将符合 /au 路径筛选条件的链接加入队列
                new_links_added = 0
                filtered_links_added = 0
                
                for link in links:
                    if link not in self.visited_urls and link not in url_queue:
                        new_links_added += 1
                        # 检查链接是否符合 /au 路径筛选条件
                        if not self.path_filter or self.matches_path_filter(link):
                            url_queue.append(link)
                            filtered_links_added += 1
                
                logger.info(f"🔗 发现 {new_links_added} 个新链接")
                logger.info(f"✅ 其中 {filtered_links_added} 个符合 /au 路径条件，已加入队列")
                
                if self.path_filter and filtered_links_added < new_links_added:
                    skipped = new_links_added - filtered_links_added
                    logger.info(f"⏭️  跳过 {skipped} 个不在 /au 路径下的链接")
                
                if self.delay > 0:
                    time.sleep(self.delay)
            else:
                logger.info(f"  ⚠️  页面状态: {status}")
    
    def _get_start_url(self):
        """获取起始URL"""
        return self.base_url
    
    def _print_crawl_info(self, start_url):
        """打印爬取信息"""
        logger.info(f"\n🚀 开始404链接检测")
        logger.info(f"🌐 目标域名: {self.domain}")
        logger.info(f"📄 最大页面数: {self.max_pages}")
        logger.info(f"🔄 并发线程数: {self.max_workers}")
        if self.path_filter:
            logger.info(f"📁 路径筛选: {self.path_filter}")
            logger.info(f"💡 策略: 首页总是被处理以获取链接，然后对发现的链接应用筛选")
        else:
            logger.info(f"📁 路径筛选: 无（检测所有页面）")
        logger.info(f"⏱️  请求延迟: {self.delay}秒")
        logger.info(f"🎯 起始URL: {start_url}")
        
        # 测试起始URL的可访问性
        logger.info(f"🔍 测试起始URL可访问性...")
        test_status = self.check_url_status(start_url)
        logger.info(f"📊 起始URL状态: {test_status}")
        
        if test_status == 'ERROR':
            logger.warning(f"⚠️  起始URL无法访问，请检查网络连接或域名是否正确")
        
        logger.info("=" * 80)
    
    def _handle_404_page(self, url):
        """处理404页面"""
        logger.info(f"  ❌ 页面本身就是404: {url}")
        with self._lock:
            self.found_404s.append({
                'url': url,
                'found_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'status_code': 404,
                'parent_page': 'N/A',
                'matches_filter': self.matches_path_filter(url),
                'position': '页面本身',
                'link_text': '',
                'element_type': 'page',
                'classes_info': [],
                'element_id': '',
                'element_tag': '',
                'css_selector': '',
                'xpath': '',
                'visual_position': '页面本身',
                'fix_suggestion': '检查页面是否已删除或URL是否正确'
            })
    
    def save_results_to_excel(self):
        """保存结果到Excel文件"""
        try:
            # 创建工作簿
            wb = Workbook()
            
            # 404链接汇总表
            ws_summary = wb.active
            ws_summary.title = "404链接汇总"
            
            # 设置表头
            headers = [
                '404链接', '发现时间', '状态码', '来源页面', '匹配筛选条件', 
                '可视化位置', 'CSS选择器', 'XPath路径', '链接文本', '元素类型', 
                '元素标签', '元素ID', 'Class信息', '修复建议'
            ]
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # 填充404链接数据
            for row, link_404 in enumerate(self.found_404s, 2):
                ws_summary.cell(row=row, column=1, value=link_404['url'])
                ws_summary.cell(row=row, column=2, value=link_404['found_time'])
                ws_summary.cell(row=row, column=3, value=link_404['status_code'])
                ws_summary.cell(row=row, column=4, value=link_404['parent_page'])
                ws_summary.cell(row=row, column=5, value='是' if link_404['matches_filter'] else '否')
                ws_summary.cell(row=row, column=6, value=link_404.get('visual_position', ''))
                ws_summary.cell(row=row, column=7, value=link_404.get('css_selector', ''))
                ws_summary.cell(row=row, column=8, value=link_404.get('xpath', ''))
                ws_summary.cell(row=row, column=9, value=link_404['link_text'])
                ws_summary.cell(row=row, column=10, value=link_404['element_type'])
                ws_summary.cell(row=row, column=11, value=link_404['element_tag'])
                ws_summary.cell(row=row, column=12, value=link_404['element_id'])
                
                # 格式化class信息
                class_info_str = ''
                for class_info in link_404.get('classes_info', []):
                    class_info_str += f"{class_info['level']}: <{class_info['tag']}> class=\"{class_info['classes']}\"\n"
                ws_summary.cell(row=row, column=13, value=class_info_str.strip())
                
                ws_summary.cell(row=row, column=14, value=link_404.get('fix_suggestion', ''))
            
            # 设置列宽
            column_widths = [50, 20, 10, 50, 15, 25, 40, 40, 30, 15, 15, 20, 60, 40]
            for col, width in enumerate(column_widths, 1):
                ws_summary.column_dimensions[ws_summary.cell(row=1, column=col).column_letter].width = width
            
            # 页面链接详情工作表
            ws_details = wb.create_sheet("页面链接详情")
            detail_headers = [
                '页面URL', '链接URL', '状态码', '检查时间', '匹配筛选条件', 
                '可视化位置', 'CSS选择器', 'XPath路径', '链接文本', '元素类型', 
                '元素标签', '元素ID', 'Class信息'
            ]
            
            # 设置详情表头样式
            for col, header in enumerate(detail_headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            detail_row = 2
            for page_detail in self.page_link_details:
                for link_status in page_detail['links_status']:
                    ws_details.cell(row=detail_row, column=1, value=page_detail['page_url'])
                    ws_details.cell(row=detail_row, column=2, value=link_status['link_url'])
                    ws_details.cell(row=detail_row, column=3, value=link_status['status_code'])
                    ws_details.cell(row=detail_row, column=4, value=link_status['check_time'])
                    ws_details.cell(row=detail_row, column=5, value='是' if link_status['matches_filter'] else '否')
                    ws_details.cell(row=detail_row, column=6, value=link_status.get('visual_position', ''))
                    ws_details.cell(row=detail_row, column=7, value=link_status.get('css_selector', ''))
                    ws_details.cell(row=detail_row, column=8, value=link_status.get('xpath', ''))
                    ws_details.cell(row=detail_row, column=9, value=link_status['link_text'])
                    ws_details.cell(row=detail_row, column=10, value=link_status['element_type'])
                    ws_details.cell(row=detail_row, column=11, value=link_status.get('element_tag', ''))
                    ws_details.cell(row=detail_row, column=12, value=link_status.get('element_id', ''))
                    
                    # 格式化class信息
                    class_info_str = ''
                    for class_info in link_status.get('classes_info', []):
                        class_info_str += f"{class_info['level']}: <{class_info['tag']}> class=\"{class_info['classes']}\"\n"
                    ws_details.cell(row=detail_row, column=13, value=class_info_str.strip())
                    
                    detail_row += 1
            
            # 设置详情表列宽
            detail_column_widths = [50, 50, 10, 20, 15, 25, 40, 40, 30, 15, 15, 20, 60]
            for col, width in enumerate(detail_column_widths, 1):
                ws_details.column_dimensions[ws_details.cell(row=1, column=col).column_letter].width = width
            
            # 统计信息工作表
            ws_stats = wb.create_sheet("统计信息")
            stats_data = [
                ['检测域名', self.domain],
                ['检测时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['已访问页面数', len(self.visited_urls)],
                ['发现链接总数', len(self.all_links)],
                ['404链接数量', len(self.found_404s)],
                ['路径筛选条件', str(self.path_filter) if self.path_filter else '无'],
            ]
            
            if self.path_filter:
                filtered_404s = [link for link in self.found_404s if link['matches_filter']]
                stats_data.append(['符合筛选条件的404链接', len(filtered_404s)])
            
            for row, (key, value) in enumerate(stats_data, 1):
                ws_stats.cell(row=row, column=1, value=key)
                ws_stats.cell(row=row, column=2, value=value)
            
            ws_stats.column_dimensions['A'].width = 25
            ws_stats.column_dimensions['B'].width = 50
            
            # 保存Excel文件
            domain_safe = self.domain.replace('.', '_').replace('://', '_')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"404_links_{domain_safe}_{timestamp}.xlsx"
            
            wb.save(filename)
            logger.info(f"\n📊 Excel报告已生成: {filename}")
            
            return filename
            
        except Exception as e:
            logger.error(f"❌ 保存Excel文件时出错: {e}")
            return None
    
    def generate_html_report(self):
        """生成HTML格式的报告"""
        try:
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>404链接检测报告 - {self.domain}</title>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                    .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    h1 {{ color: #333; text-align: center; border-bottom: 3px solid #007bff; padding-bottom: 10px; }}
                    .summary {{ background: #e9ecef; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                    .link-item {{ border: 1px solid #ddd; margin: 15px 0; padding: 20px; border-radius: 8px; background: #fff; }}
                    .link-item:hover {{ box-shadow: 0 4px 8px rgba(0,0,0,0.1); }}
                    .url {{ color: #d32f2f; font-weight: bold; font-size: 16px; word-break: break-all; }}
                    .position {{ color: #1976d2; background: #e3f2fd; padding: 5px 10px; border-radius: 3px; display: inline-block; margin: 5px 0; }}
                    .selector {{ background: #f8f9fa; padding: 10px; font-family: 'Courier New', monospace; border-left: 4px solid #007bff; margin: 10px 0; word-break: break-all; }}
                    .suggestion {{ background: #fff3cd; border: 1px solid #ffeaa7; padding: 10px; border-radius: 5px; margin: 10px 0; }}
                    .meta {{ color: #666; font-size: 14px; }}
                    .tag {{ background: #6c757d; color: white; padding: 2px 6px; border-radius: 3px; font-size: 12px; margin: 2px; }}
                    .stats {{ display: flex; justify-content: space-around; margin: 20px 0; }}
                    .stat-item {{ text-align: center; padding: 15px; background: #f8f9fa; border-radius: 5px; }}
                    .stat-number {{ font-size: 24px; font-weight: bold; color: #007bff; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>🔍 404链接检测报告</h1>
                    
                    <div class="summary">
                        <p><strong>🌐 检测域名:</strong> {self.domain}</p>
                        <p><strong>⏰ 检测时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    </div>
                    
                    <div class="stats">
                        <div class="stat-item">
                            <div class="stat-number">{len(self.visited_urls)}</div>
                            <div>已检测页面</div>
                        </div>
                        <div class="stat-item">
                            <div class="stat-number">{len(self.all_links)}</div>
                            <div>发现链接总数</div>
                        </div>
                        <div class="stat-item">
                            <div class="stat-number">{len(self.found_404s)}</div>
                            <div>404链接数量</div>
                        </div>
                        <div class="stat-item">
                            <div class="stat-number">{sum(1 for link in self.found_404s if link['matches_filter'])}</div>
                            <div>符合筛选条件</div>
                        </div>
                    </div>
            """
            
            for i, link_404 in enumerate(self.found_404s, 1):
                html_content += f"""
                <div class="link-item">
                    <h3>#{i} <span class="url">{link_404['url']}</span></h3>
                    <div class="meta">
                        <p><strong>📄 来源页面:</strong> {link_404['parent_page']}</p>
                        <p><strong>⏰ 发现时间:</strong> {link_404['found_time']}</p>
                        <p><strong>📍 位置:</strong> <span class="position">{link_404.get('visual_position', '未知')}</span></p>
                        {f'<p><strong>📝 链接文本:</strong> {link_404["link_text"]}</p>' if link_404.get('link_text') else ''}
                        <p><strong>🏷️ 元素类型:</strong> <span class="tag">{link_404['element_type']}</span> 
                           <span class="tag">{link_404['element_tag']}</span>
                           {f'<span class="tag">ID: {link_404["element_id"]}</span>' if link_404.get('element_id') else ''}</p>
                    </div>
                    
                    <div class="selector">
                        <strong>🎯 CSS选择器:</strong><br>
                        <code>{link_404.get('css_selector', '未生成')}</code>
                    </div>
                    
                    <div class="selector">
                        <strong>🗺️ XPath路径:</strong><br>
                        <code>{link_404.get('xpath', '未生成')}</code>
                    </div>
                    
                    <div class="suggestion">
                        <strong>💡 修复建议:</strong> {link_404.get('fix_suggestion', '手动检查链接有效性')}
                    </div>
                </div>
                """
            
            html_content += """
                </div>
            </body>
            </html>
            """
            
            # 保存HTML文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"404_report_{self.domain.replace('.', '_')}_{timestamp}.html"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"📄 HTML报告已保存: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"生成HTML报告时出错: {e}")
            return None
    
    def save_json_report(self):
        """保存JSON格式的详细报告"""
        try:
            report_data = {
                'scan_info': {
                    'domain': self.domain,
                    'scan_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'max_pages': self.max_pages,
                    'path_filter': self.path_filter,
                    'total_pages_scanned': len(self.visited_urls),
                    'total_links_found': len(self.all_links),
                    'total_404s_found': len(self.found_404s)
                },
                'found_404s': self.found_404s,
                'page_details': self.page_link_details
            }
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"404_data_{self.domain.replace('.', '_')}_{timestamp}.json"
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"📋 JSON数据已保存: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"保存JSON文件时出错: {e}")
            return None
    
    def cleanup(self):
        """清理资源"""
        try:
            if hasattr(self, 'session'):
                self.session.close()
            logger.info("🧹 资源清理完成")
        except Exception as e:
            logger.error(f"清理资源时出错: {e}")
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.cleanup()
    
    def _generate_css_selector(self, element):
        """生成CSS选择器"""
        try:
            selectors = []
            current = element
            
            while current and current.name and len(selectors) < 5:
                selector_part = current.name
                
                # 添加ID
                if current.get('id'):
                    selector_part += f"#{current['id']}"
                    selectors.append(selector_part)
                    break  # ID是唯一的，可以停止
                
                # 添加class
                classes = current.get('class', [])
                if classes:
                    # 只使用前两个class以避免选择器过长
                    class_str = '.'.join(classes[:2])
                    selector_part += f".{class_str}"
                
                selectors.append(selector_part)
                current = current.parent
            
            return ' > '.join(reversed(selectors))
            
        except Exception:
            return ''
    
    def _generate_xpath(self, element):
        """生成XPath路径"""
        try:
            xpath_parts = []
            current = element
            
            while current and current.name and len(xpath_parts) < 5:
                tag = current.name
                
                # 如果有ID，使用ID定位
                if current.get('id'):
                    xpath_parts.append(f"//{tag}[@id='{current['id']}']")
                    break
                
                # 计算同级元素中的位置
                siblings = [s for s in current.parent.children if hasattr(s, 'name') and s.name == tag] if current.parent else []
                if len(siblings) > 1:
                    index = siblings.index(current) + 1
                    xpath_parts.append(f"{tag}[{index}]")
                else:
                    xpath_parts.append(tag)
                
                current = current.parent
            
            if xpath_parts:
                return '/' + '/'.join(reversed(xpath_parts))
            return ''
            
        except Exception:
            return ''
    
    def _determine_visual_position(self, positions, element_classes):
        """确定元素的可视化位置"""
        # 基于位置信息确定可视化区域
        position_str = ' '.join(positions).lower()
        class_str = ' '.join(element_classes).lower() if element_classes else ''
        
        combined = f"{position_str} {class_str}"
        
        if any(keyword in combined for keyword in ['header', 'top', 'navbar', 'nav-bar']):
            return '页面头部'
        elif any(keyword in combined for keyword in ['footer', 'bottom', 'foot']):
            return '页面底部'
        elif any(keyword in combined for keyword in ['sidebar', 'side-bar', 'aside']):
            return '侧边栏'
        elif any(keyword in combined for keyword in ['menu', 'nav', 'navigation']):
            return '导航菜单'
        elif any(keyword in combined for keyword in ['breadcrumb']):
            return '面包屑导航'
        elif any(keyword in combined for keyword in ['pagination', 'pager']):
            return '分页区域'
        elif any(keyword in combined for keyword in ['search']):
            return '搜索区域'
        elif any(keyword in combined for keyword in ['social', 'share']):
            return '社交分享区域'
        elif any(keyword in combined for keyword in ['product', 'item', 'card']):
            return '产品/内容卡片'
        elif any(keyword in combined for keyword in ['category', 'section']):
            return '分类/栏目区域'
        elif any(keyword in combined for keyword in ['banner', 'hero', 'slider']):
            return '横幅/轮播区域'
        elif any(keyword in combined for keyword in ['form', 'contact']):
            return '表单区域'
        elif any(keyword in combined for keyword in ['main', 'content', 'body']):
            return '主要内容区域'
        else:
            return '页面主体'
    
    def generate_fix_suggestion(self, link_info):
        """生成修复建议"""
        url = link_info.get('url', '')
        position = link_info.get('visual_position', '')
        
        suggestions = []
        
        # 基于URL模式的建议
        if '/product' in url.lower():
            suggestions.append('检查产品是否已下架或ID是否正确')
        elif '/category' in url.lower():
            suggestions.append('检查分类是否已删除或重命名')
        elif '/blog' in url.lower() or '/post' in url.lower():
            suggestions.append('检查文章是否已删除或URL结构是否变更')
        elif '/user' in url.lower() or '/profile' in url.lower():
            suggestions.append('检查用户账户是否存在或权限设置')
        
        # 基于位置的建议
        if '导航' in position:
            suggestions.append('更新导航菜单配置')
        elif '底部' in position:
            suggestions.append('检查页脚链接配置')
        elif '侧边栏' in position:
            suggestions.append('更新侧边栏组件')
        
        # 通用建议
        suggestions.extend([
            '检查目标页面是否存在',
            '验证URL拼写是否正确',
            '确认链接目标是否已迁移'
        ])
        
        return '; '.join(suggestions[:3])  # 返回前3个建议

def get_user_config():
    """获取用户配置"""
    print("\n🔧 404链接检测工具配置")
    print("=" * 50)
    
    # 获取目标域名
    while True:
        domain = input("\n🌐 请输入目标域名 (例如: www.anker.com): ").strip()
        if domain:
            break
        print("❌ 域名不能为空，请重新输入")
    
    # 获取路径筛选
    print("\n📁 选择检测范围:")
    print("1. 检测所有页面 (无筛选)")
    print("2. 只检测澳洲站点 (/au)")
    print("3. 只检测产品页面 (/products)")
    print("4. 只检测支持页面 (/support)")
    print("5. 只检测博客页面 (/blog)")
    print("6. 自定义路径筛选")
    print("7. 多路径筛选")
    
    while True:
        try:
            filter_choice = int(input("\n请选择 (1-7): ").strip())
            if 1 <= filter_choice <= 7:
                break
            print("❌ 请输入1-7之间的数字")
        except ValueError:
            print("❌ 请输入有效的数字")
    
    path_filter = None
    if filter_choice == 2:
        path_filter = '/au'
        print("✅ 已选择澳洲站点，将只检测 /au 路径下的页面")
    elif filter_choice == 3:
        path_filter = '/products'
    elif filter_choice == 4:
        path_filter = '/support'
    elif filter_choice == 5:
        path_filter = '/blog'
    elif filter_choice == 6:
        custom_path = input("\n请输入自定义路径 (例如: /au, /category): ").strip()
        if custom_path:
            if not custom_path.startswith('/'):
                custom_path = '/' + custom_path
            path_filter = custom_path
    elif filter_choice == 7:
        print("\n请输入多个路径，用逗号分隔 (例如: /au,/products,/support):")
        multi_paths = input().strip()
        if multi_paths:
            paths = [p.strip() for p in multi_paths.split(',')]
            paths = [p if p.startswith('/') else '/' + p for p in paths if p]
            if paths:
                path_filter = paths
    
    # 获取最大检测页面数
    while True:
        try:
            max_pages = int(input("\n📄 最大检测页面数 (建议10-100，默认50): ").strip() or "50")
            if max_pages > 0:
                break
            print("❌ 页面数必须大于0")
        except ValueError:
            print("❌ 请输入有效的数字")
    
    # 获取并发线程数
    while True:
        try:
            max_workers = int(input("\n🔄 并发线程数 (建议3-10，默认5): ").strip() or "5")
            if 1 <= max_workers <= 20:
                break
            print("❌ 线程数建议在1-20之间")
        except ValueError:
            print("❌ 请输入有效的数字")
    
    # 获取请求延迟
    while True:
        try:
            delay = float(input("\n⏱️ 请求延迟秒数 (建议0.5-2，默认1): ").strip() or "1")
            if delay >= 0:
                break
            print("❌ 延迟时间不能为负数")
        except ValueError:
            print("❌ 请输入有效的数字")
    
    # 确认配置
    print("\n📋 配置确认:")
    print(f"🌐 目标域名: {domain}")
    if isinstance(path_filter, list):
        print(f"📁 检测路径: {', '.join(path_filter)}")
    else:
        print(f"📁 检测路径: {path_filter if path_filter else '所有页面'}")
    print(f"📄 最大页面数: {max_pages}")
    print(f"🔄 并发线程数: {max_workers}")
    print(f"⏱️ 请求延迟: {delay}秒")
    
    confirm = input("\n✅ 确认开始检测？(y/n，默认y): ").strip().lower()
    if confirm in ['n', 'no']:
        print("❌ 用户取消检测")
        return None
    
    return {
        'domain': domain,
        'path_filter': path_filter,
        'max_pages': max_pages,
        'max_workers': max_workers,
        'delay': delay
    }

def main():
    """主函数"""
    try:
        # 获取用户配置
        config = get_user_config()
        if not config:
            return
        
        # 创建爬虫实例并开始检测
        with Link404Crawler(
            domain=config['domain'],
            max_pages=config['max_pages'],
            delay=config['delay'],
            path_filter=config['path_filter'],
            max_workers=config['max_workers']
        ) as crawler:
            
            # 开始爬取
            crawler.crawl_for_404s()
            
            # 打印最终摘要
            crawler.print_final_summary()
            
            # 修复：无论是否发现404都保存报告
            print("\n💾 正在保存检测结果...")
            
            # 保存Excel报告
            excel_file = crawler.save_results_to_excel()
            
            # 保存HTML报告
            html_file = crawler.generate_html_report()
            
            # 保存JSON数据
            json_file = crawler.save_json_report()
            
            print("\n🎉 检测完成！生成的文件:")
            if excel_file:
                print(f"📊 Excel报告: {excel_file}")
            if html_file:
                print(f"📄 HTML报告: {html_file}")
            if json_file:
                print(f"📋 JSON数据: {json_file}")
                
            if not crawler.found_404s:
                print("\n🎉 太棒了！没有发现404链接！")
                
    except KeyboardInterrupt:
        print("\n\n⚠️ 用户中断检测")
    except Exception as e:
        logger.error(f"程序执行出错: {e}")
        print(f"\n❌ 程序执行出错: {e}")

if __name__ == "__main__":
    main()