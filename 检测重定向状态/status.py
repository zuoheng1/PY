import requests
import openpyxl
from openpyxl import Workbook

# 从Excel文件读取URL
def load_urls_from_excel(file_path):
    urls = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # 使用活动工作表
        
        # 读取第一列的所有URL（跳过标题行如果有的话）
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value:
                # 清理URL并添加前缀
                url_path = str(cell_value).strip()
                # 如果已经是完整URL，提取路径部分
                if url_path.startswith('http'):
                    from urllib.parse import urlparse
                    parsed = urlparse(url_path)
                    url_path = parsed.path
                
                # 确保路径以/开头
                if not url_path.startswith('/'):
                    url_path = '/' + url_path
                
                # 添加www.anker.com前缀
                full_url = 'https://www.anker.com' + url_path
                urls.append(full_url)
        
        workbook.close()
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
    
    return urls

# 检查URL重定向信息
def check_url_redirect(url):
    try:
        # 不跟随重定向，获取初始响应
        response = requests.head(url, allow_redirects=False, timeout=10)
        
        # 如果是重定向状态码
        if response.status_code in [301, 302, 303, 307, 308]:
            # 获取重定向目标
            redirect_url = response.headers.get('Location', '')
            
            # 处理相对URL
            if redirect_url.startswith('/'):
                from urllib.parse import urljoin
                redirect_url = urljoin(url, redirect_url)
            
            return {
                'original_url': url,
                'redirect_url': redirect_url,
                'status_code': response.status_code
            }
        else:
            # 没有重定向
            return {
                'original_url': url,
                'redirect_url': url,  # 没有重定向，最终URL就是原始URL
                'status_code': response.status_code
            }
    except Exception as e:
        return {
            'original_url': url,
            'redirect_url': f'ERROR: {e}',
            'status_code': 'ERROR'
        }

# 保存结果到Excel
def save_results_to_excel(results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "重定向检查结果"
    
    # 设置表头
    ws['A1'] = '原始URL'
    ws['B1'] = '重定向后URL'
    ws['C1'] = '状态码'
    
    # 写入数据
    for i, result in enumerate(results, start=2):
        ws[f'A{i}'] = result['original_url']
        ws[f'B{i}'] = result['redirect_url']
        ws[f'C{i}'] = result['status_code']
    
    # 调整列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    
    wb.save(output_file)
    print(f"结果已保存到: {output_file}")

# 主程序
if __name__ == "__main__":
    # 从Excel文件加载URL
    urls = load_urls_from_excel('url.xlsx')
    
    if not urls:
        print("没有找到有效的URL")
    else:
        print(f"找到 {len(urls)} 个URL，开始检查重定向...")
        
        results = []
        
        for i, url in enumerate(urls, 1):
            print(f"检查 {i}/{len(urls)}: {url}")
            result = check_url_redirect(url)
            results.append(result)
            
            # 显示结果
            if result['status_code'] in [301, 302, 303, 307, 308]:
                print(f"  重定向: {result['status_code']} -> {result['redirect_url']}")
            elif result['status_code'] == 'ERROR':
                print(f"  错误: {result['redirect_url']}")
            else:
                print(f"  正常: {result['status_code']}")
        
        # 保存结果到Excel文件
        output_file = 'redirect_results.xlsx'
        save_results_to_excel(results, output_file)
        
        print(f"\n检查完成！结果已保存到 {output_file}")
        print(f"共检查了 {len(results)} 个URL")
        redirect_count = sum(1 for r in results if r['status_code'] in [301, 302, 303, 307, 308])
        print(f"其中 {redirect_count} 个URL有重定向")
